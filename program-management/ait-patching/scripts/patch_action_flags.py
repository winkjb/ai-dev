"""
AIT patching - action flags for account management (one-off, 2026-07-21).

Reads the monthly "Patch Status" export and finds every device sitting in
Patch Status Failed or Not Installed - those are the ones account management
needs to open a ticket against. Two nuances applied before a device counts
as actionable:

  - Windows 10 is end of life and can no longer be patched, so a Windows 10
    device in Failed/Not Installed is expected noise, not a real gap - unless
    the customer has an ESU (Extended Security Updates) agreement, in which
    case patches are still expected and the flag is real. ESU customers are
    listed in ../data/reference/customers-with-esu.csv.
  - OS comes from the Network Hardware sheet (device inventory), joined on
    Customer + Device Name - the per-customer detail sheets don't carry OS
    themselves.
  - Customers in ../data/reference/customers-to-ignore.csv are dropped
    entirely (account management doesn't track patch status for them at
    all, regardless of device status).
  - Only Windows Laptop and Windows Workstation are in scope - account
    management doesn't action servers here. Windows Server rows are
    dropped (the only three Device Class values that appear in this
    workbook's patch-status sheets are Laptop/Workstation/Server).

Output is grouped by Customer + Site (location), since account management
wants to open one ticket per location covering every flagged machine there.

The source workbook has one sheet per customer (132 of them, inconsistently
named Sheet6..Sheet137) plus a handful of intro/summary sheets (Document map,
Patch Summary, Patch Status by Customer, Network Hardware). Rather than rely
on sheet names or the Document map's hyperlink order (fragile - shifts if
sheets are added/removed next month), every sheet is scanned for a cell
starting with "Customer:" to detect customer-detail sheets directly.

Usage:
    python patch_action_flags.py

Reads:  ../data/raw/July Patch Status.xlsx
        ../data/reference/customers-with-esu.csv
        ../data/reference/customers-to-ignore.csv
Writes: ../data/reports/patch-action-flags-detail.csv       (one row per flagged device/patch category)
        ../data/reports/patch-action-flags-by-location.csv  (one row per flagged device, ticket-ready, grouped by location)
        ../data/reports/patch-action-flags-summary.md
"""

import csv
from pathlib import Path

import openpyxl

# --- config -------------------------------------------------------------

RAW_WORKBOOK = Path("../data/raw/July Patch Status.xlsx")
ESU_LIST = Path("../data/reference/customers-with-esu.csv")
IGNORE_LIST = Path("../data/reference/customers-to-ignore.csv")
OUTPUT_DIR = Path("../data/reports")
OUTPUT_DETAIL = OUTPUT_DIR / "patch-action-flags-detail.csv"
OUTPUT_BY_LOCATION = OUTPUT_DIR / "patch-action-flags-by-location.csv"
OUTPUT_SUMMARY = OUTPUT_DIR / "patch-action-flags-summary.md"

ACTIONABLE_STATUSES = {"Failed", "Not Installed"}
IN_SCOPE_DEVICE_CLASSES = {"Windows Laptop", "Windows Workstation"}


def load_esu_customers():
    with open(ESU_LIST, encoding="utf-8-sig") as f:
        return {row["Customer"].strip() for row in csv.DictReader(f)}


def load_ignored_customers():
    with open(IGNORE_LIST, encoding="utf-8-sig") as f:
        return {row["CustomerName"].strip() for row in csv.DictReader(f)}


def load_hardware_lookup(wb):
    """(Customer, Device Name) -> {"site": ..., "os": ...} from Network Hardware."""
    ws = wb["Network Hardware"]
    lookup = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        customer, site, device_name, os_val = row[1], row[2], row[4], row[10]
        if not customer or not device_name:
            continue
        lookup[(customer, device_name)] = {"site": site or "No Site", "os": os_val}
    return lookup


def find_customer_detail_sheets(wb):
    """Sheet name -> customer name, for every sheet with a 'Customer: X' cell."""
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 5:
                break
            for cell in row:
                if isinstance(cell, str) and cell.startswith("Customer:"):
                    sheets[name] = cell.split("Customer:", 1)[1].strip()
                    break
            if name in sheets:
                break
    return sheets


def parse_customer_sheet(ws, customer):
    """Yield (status, device_class, device_name, patch_category, count) for
    every data row, walking the Patch Status sub-sections top to bottom."""
    status = None
    device_class = None
    device_name = None
    for row in ws.iter_rows(values_only=True):
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        e = row[4] if len(row) > 4 else None
        f = row[5] if len(row) > 5 else None

        if isinstance(b, str) and b.startswith("Customer:"):
            continue
        if isinstance(c, str) and c.startswith("Patch Status:"):
            status = c.split("Patch Status:", 1)[1].strip()
            device_class, device_name = None, None
            continue
        if c == "Device Class":
            continue
        if isinstance(c, str) and c.strip().endswith("Patch Count:"):
            continue
        if e:
            if c:
                device_class = c
            if d:
                device_name = d
            if device_name:
                yield status, device_class, device_name, e, f


def build_flags(wb, esu_customers, ignored_customers):
    hw_lookup = load_hardware_lookup(wb)
    detail_sheets = find_customer_detail_sheets(wb)

    records = []
    win10_excluded = 0
    server_excluded = 0
    unmatched = 0
    ignored_customers_seen = set()

    for sheet_name, customer in detail_sheets.items():
        if customer in ignored_customers:
            ignored_customers_seen.add(customer)
            continue
        ws = wb[sheet_name]
        for status, device_class, device_name, patch_category, count in parse_customer_sheet(ws, customer):
            if status not in ACTIONABLE_STATUSES:
                continue
            if device_class not in IN_SCOPE_DEVICE_CLASSES:
                server_excluded += 1
                continue

            hw = hw_lookup.get((customer, device_name))
            matched = hw is not None
            os_val = hw["os"] if hw else None
            site = hw["site"] if hw else "Unknown (not in Network Hardware)"

            is_win10 = bool(os_val) and os_val.strip().startswith("10 ")
            has_esu = customer in esu_customers

            if is_win10 and not has_esu:
                win10_excluded += 1
                continue
            if not matched:
                unmatched += 1

            records.append({
                "Customer": customer,
                "Site": site,
                "Device Class": device_class,
                "Device Name": device_name,
                "OS": os_val or "Unknown",
                "Status": status,
                "Patch Category": patch_category,
                "Count": count,
                "Matched to Inventory": matched,
            })

    return records, win10_excluded, server_excluded, unmatched, ignored_customers_seen


def write_detail(records):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    cols = ["Customer", "Site", "Device Class", "Device Name", "OS", "Status",
            "Patch Category", "Count", "Matched to Inventory"]
    ordered = sorted(records, key=lambda r: (r["Customer"], r["Site"], r["Device Name"], r["Status"]))
    with open(OUTPUT_DETAIL, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        writer.writerows(ordered)


def write_by_location(records):
    """One row per flagged device (deduped across patch categories), grouped
    by Customer + Site so account management can open one ticket per group."""
    devices = {}
    for r in records:
        key = (r["Customer"], r["Site"], r["Device Name"])
        d = devices.setdefault(key, {
            "Customer": r["Customer"], "Site": r["Site"],
            "Device Class": r["Device Class"], "Device Name": r["Device Name"],
            "OS": r["OS"], "Statuses": set(), "Matched to Inventory": r["Matched to Inventory"],
        })
        d["Statuses"].add(r["Status"])

    cols = ["Customer", "Site", "Device Class", "Device Name", "OS", "Status", "Matched to Inventory"]
    ordered = sorted(devices.values(), key=lambda d: (d["Customer"], d["Site"], d["Device Name"]))
    with open(OUTPUT_BY_LOCATION, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        for d in ordered:
            writer.writerow({
                "Customer": d["Customer"], "Site": d["Site"],
                "Device Class": d["Device Class"], "Device Name": d["Device Name"],
                "OS": d["OS"], "Status": ", ".join(sorted(d["Statuses"])),
                "Matched to Inventory": d["Matched to Inventory"],
            })
    return ordered


def write_summary(records, by_location, win10_excluded, server_excluded, unmatched,
                   esu_customers, ignored_customers_seen):
    unique_devices = {(r["Customer"], r["Site"], r["Device Name"]) for r in records}
    locations = {(r["Customer"], r["Site"]) for r in records}

    location_counts = {}
    for d in by_location:
        key = (d["Customer"], d["Site"])
        location_counts[key] = location_counts.get(key, 0) + 1
    top_locations = sorted(location_counts.items(), key=lambda kv: kv[1], reverse=True)

    lines = []
    lines.append("# AIT Patch Action Flags - Summary")
    lines.append("")
    lines.append("Workstations and laptops (no servers) in Patch Status Failed or Not "
                 "Installed, excluding unpatchable Windows 10 devices (EOL) unless the "
                 "customer has an ESU agreement, and excluding customers on the ignore list.")
    lines.append("")
    lines.append(f"- Flagged devices (unique, across all locations): {len(unique_devices)}")
    lines.append(f"- Locations needing a ticket (Customer + Site): {len(locations)}")
    lines.append(f"- Windows 10 rows excluded as unpatchable (no ESU): {win10_excluded}")
    lines.append(f"- Windows Server rows excluded (out of scope - workstations/laptops only): {server_excluded}")
    lines.append(f"- Flagged rows with no OS match in Network Hardware (kept, flagged as unmatched): {unmatched}")
    lines.append(f"- Customers currently on the ESU list: {', '.join(sorted(esu_customers)) or 'none'}")
    lines.append(f"- Customers skipped entirely (ignore list): {', '.join(sorted(ignored_customers_seen)) or 'none'}")
    lines.append("")
    lines.append("## Locations by device count (top 20)")
    lines.append("")
    lines.append("| Customer | Site | Flagged Devices |")
    lines.append("|---|---|---|")
    for (customer, site), count in top_locations[:20]:
        customer_display = customer.replace("|", "\\|")
        site_display = site.replace("|", "\\|")
        lines.append(f"| {customer_display} | {site_display} | {count} |")
    lines.append("")
    lines.append(f"Ticket-ready detail (one row per device, grouped by location): {OUTPUT_BY_LOCATION.name}  ")
    lines.append(f"Full detail (one row per flagged patch category): {OUTPUT_DETAIL.name}")

    OUTPUT_SUMMARY.write_text("\n".join(lines))


def main():
    wb = openpyxl.load_workbook(RAW_WORKBOOK, data_only=True, read_only=True)
    esu_customers = load_esu_customers()
    ignored_customers = load_ignored_customers()

    records, win10_excluded, server_excluded, unmatched, ignored_customers_seen = build_flags(
        wb, esu_customers, ignored_customers
    )
    write_detail(records)
    by_location = write_by_location(records)
    write_summary(records, by_location, win10_excluded, server_excluded, unmatched,
                  esu_customers, ignored_customers_seen)

    unique_devices = {(r["Customer"], r["Site"], r["Device Name"]) for r in records}
    locations = {(r["Customer"], r["Site"]) for r in records}
    print(f"Flagged devices: {len(unique_devices)} across {len(locations)} location(s)")
    print(f"Windows 10 (unpatchable, no ESU) excluded: {win10_excluded}")
    print(f"Windows Server rows excluded (out of scope): {server_excluded}")
    print(f"Flagged rows unmatched in Network Hardware: {unmatched}")
    print(f"Customers skipped (ignore list): {sorted(ignored_customers_seen)}")
    print(f"Wrote {OUTPUT_DETAIL}, {OUTPUT_BY_LOCATION}, and {OUTPUT_SUMMARY}")


if __name__ == "__main__":
    main()
